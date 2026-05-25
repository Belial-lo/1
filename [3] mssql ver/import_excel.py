from pathlib import Path

import pyodbc
from openpyxl import load_workbook


BASE_DIR = Path(__file__).parent


connection_string = (
    "DRIVER={SQL Server};"
    "SERVER=localhost\\SQLEXPRESS;"
    "DATABASE=DB_Ivah;"
    "Trusted_Connection=yes;"
    "Encrypt=no;"
    "TrustServerCertificate=yes;"
)

conn = pyodbc.connect(connection_string)
cursor = conn.cursor()


# 1. Импорт пользователей
cursor.execute("IF NOT EXISTS (SELECT 1 FROM roles WHERE name = N'admin') INSERT INTO roles (name) VALUES (N'admin')")
cursor.execute("IF NOT EXISTS (SELECT 1 FROM roles WHERE name = N'manager') INSERT INTO roles (name) VALUES (N'manager')")
cursor.execute("IF NOT EXISTS (SELECT 1 FROM roles WHERE name = N'client') INSERT INTO roles (name) VALUES (N'client')")

workbook = load_workbook(BASE_DIR / "user_import.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    role_text = row[0]
    full_name = row[1]
    login = row[2]
    password = row[3]

    if login is None:
        continue

    if role_text == "Администратор":
        role_name = "admin"
    elif role_text == "Менеджер":
        role_name = "manager"
    else:
        role_name = "client"

    cursor.execute("SELECT id FROM roles WHERE name = ?", role_name)
    role_id = cursor.fetchone()[0]

    cursor.execute(
        """
        IF NOT EXISTS (SELECT 1 FROM users WHERE login = ?)
        INSERT INTO users (full_name, login, password, role_id)
        VALUES (?, ?, ?, ?)
        """,
        login,
        full_name,
        login,
        password,
        role_id,
    )


# 2. Импорт товаров
workbook = load_workbook(BASE_DIR / "Tovar.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    article_num = row[0]
    product = row[1]
    unit = row[2]
    price = row[3]
    supplier = row[4]
    creator = row[5]
    category = row[6]
    discount = row[7]
    stock_quantity = row[8]
    description = row[9]
    photo = row[10]

    if article_num is None:
        continue

    cursor.execute(
        """
        IF NOT EXISTS (SELECT 1 FROM products WHERE article_num = ?)
        INSERT INTO products (
            article_num, product, unit, price, supplier, creator,
            category, discount, stock_quantity, description, photo
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        article_num,
        article_num,
        product,
        unit,
        price,
        supplier,
        creator,
        category,
        discount,
            stock_quantity,
        description,
        f"image/{photo}" if photo else "",
    )


conn.commit()
cursor.close()
conn.close()

print("Импорт завершен")
