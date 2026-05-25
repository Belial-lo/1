import hashlib
from pathlib import Path

import psycopg2
from openpyxl import load_workbook


DB_CONFIG = {
    "host": "localhost",
    "database": "DB_p",
    "user": "postgres",
    "password": "admin",
}

BASE_DIR = Path(__file__).parent


def hash_password(password):
    return hashlib.sha256(str(password).encode("utf-8")).hexdigest()


conn = psycopg2.connect(**DB_CONFIG)
cursor = conn.cursor()


# 1. Роли и пользователи из user_import.xlsx
cursor.execute("INSERT INTO roles (name) VALUES ('guest') ON CONFLICT (name) DO NOTHING")
cursor.execute("INSERT INTO roles (name) VALUES ('client') ON CONFLICT (name) DO NOTHING")
cursor.execute("INSERT INTO roles (name) VALUES ('manager') ON CONFLICT (name) DO NOTHING")
cursor.execute("INSERT INTO roles (name) VALUES ('admin') ON CONFLICT (name) DO NOTHING")

workbook = load_workbook(BASE_DIR / "user_import.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    role_text = row[0]
    full_name = row[1]
    login = row[2]
    password = row[3]

    if login is None:
        continue

    if role_text == "Администратор":
        role_name = "admin"
    elif role_text == "Менеджер":
        role_name = "manager"
    else:
        role_name = "client"

    cursor.execute("SELECT id FROM roles WHERE name=%s", [role_name])
    role_id = cursor.fetchone()[0]

    cursor.execute(
        """
        INSERT INTO users (role_id, login, password_hash, full_name)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (login) DO UPDATE
        SET role_id=EXCLUDED.role_id,
            password_hash=EXCLUDED.password_hash,
            full_name=EXCLUDED.full_name
        """,
        [role_id, login, hash_password(password), full_name],
    )


# 2. Товары из Tovar.xlsx
workbook = load_workbook(BASE_DIR / "Tovar.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    article = row[0]
    name = row[1]
    unit = row[2]
    price = row[3]
    supplier = row[4]
    manufacturer = row[5]
    category = row[6]
    discount = row[7]
    stock_quantity = row[8]
    description = row[9]
    image_name = row[10]

    if article is None:
        continue

    image_path = f"image/{image_name}" if image_name else ""

    cursor.execute(
        """
        INSERT INTO products (
            article, name, category, description, manufacturer, supplier,
            price, discount, unit, stock_quantity, image_path
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (article) DO UPDATE
        SET name=EXCLUDED.name,
            category=EXCLUDED.category,
            description=EXCLUDED.description,
            manufacturer=EXCLUDED.manufacturer,
            supplier=EXCLUDED.supplier,
            price=EXCLUDED.price,
            discount=EXCLUDED.discount,
            unit=EXCLUDED.unit,
            stock_quantity=EXCLUDED.stock_quantity,
            image_path=EXCLUDED.image_path
        """,
        [
            article,
            name,
            category,
            description,
            manufacturer,
            supplier,
            price,
            discount,
            unit,
            stock_quantity,
            image_path,
        ],
    )


conn.commit()
cursor.close()
conn.close()

print("Импорт пользователей и товаров завершен")
