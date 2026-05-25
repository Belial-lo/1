from pathlib import Path

import psycopg2
from openpyxl import load_workbook


DB_CONFIG = {
    "host": "localhost",
    "database": "DB_Ivah",
    "user": "postgres",
    "password": "admin",
}

BASE_DIR = Path(__file__).parent


conn = psycopg2.connect(**DB_CONFIG)
cursor = conn.cursor()


# 1. Импорт пользователей
cursor.execute("INSERT INTO roles (name) VALUES ('admin') ON CONFLICT (name) DO NOTHING")
cursor.execute("INSERT INTO roles (name) VALUES ('manager') ON CONFLICT (name) DO NOTHING")
cursor.execute("INSERT INTO roles (name) VALUES ('client') ON CONFLICT (name) DO NOTHING")

workbook = load_workbook(BASE_DIR / "user_import.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    role_text = row[0]
    full_name = row[1]
    login = row[2]
    password = row[3]

    if login is None:
        continue

    if role_text == "Администратор":
        role_name = "admin"
    elif role_text == "Менеджер":
        role_name = "manager"
    else:
        role_name = "client"

    cursor.execute("SELECT id FROM roles WHERE name=%s", [role_name])
    role_id = cursor.fetchone()[0]

    cursor.execute(
        """
        INSERT INTO users (full_name, login, password, role_id)
        VALUES (%s, %s, %s, %s)
        ON CONFLICT (login) DO NOTHING
        """,
        [full_name, login, password, role_id],
    )


# 2. Импорт товаров
workbook = load_workbook(BASE_DIR / "Tovar.xlsx")
sheet = workbook.active

for row in sheet.iter_rows(min_row=2, values_only=True):
    article_num = row[0]
    product = row[1]
    unit = row[2]
    price = row[3]
    supplier = row[4]
    creator = row[5]
    category = row[6]
    discount = row[7]
    inventory = row[8]
    description = row[9]
    photo = row[10]

    if article_num is None:
        continue

    cursor.execute(
        """
        INSERT INTO products (
            article_num, product, unit, price, supplier, creator,
            category, discount, inventory, description, photo
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ON CONFLICT (article_num) DO NOTHING
        """,
        [
            article_num,
            product,
            unit,
            price,
            supplier,
            creator,
            category,
            discount,
            inventory,
            description,
            f"image/{photo}" if photo else "",
        ],
    )


conn.commit()
cursor.close()
conn.close()

print("Импорт завершен")
